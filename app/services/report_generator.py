"""
Generador de reportes en Excel y PDF
"""

from sqlalchemy.orm import Session, joinedload
from datetime import datetime, timedelta
from app.models.sale import Sale
from app.models.sale_item import SaleItem
from app.models.product import Product
from app.models.daily_box import DailyBox
import io

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.units import inch
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False


def _get_period_sales(db: Session, company_id: int, start_date: datetime, end_date: datetime):
    """
    Trae las ventas de una compañía en un rango de fechas, con items y
    producto precargados. No incluye ventas anuladas: mismo criterio que
    usa el resto del sistema (dashboard, caja diaria) para no contar como
    ingreso/ganancia algo que se revirtió.
    """
    return db.query(Sale).filter(
        Sale.company_id == company_id,
        Sale.created_at >= start_date,
        Sale.created_at <= end_date,
        Sale.cancelled_at.is_(None)
    ).options(
        joinedload(Sale.items).joinedload(SaleItem.product),
        joinedload(Sale.customer)
    ).order_by(Sale.created_at.desc()).all()


def _summarize_sales(sales: list) -> dict:
    """
    Calcula los totales de un período a partir de una lista de ventas ya
    filtrada (activas, de una compañía). El ingreso se toma de
    Sale.total_amount (ya refleja descuentos aplicados en el checkout); el
    costo sale de los ítems, que no cambia con el descuento.
    """
    total_revenue = 0.0
    total_cost = 0.0
    paid_count = 0
    total_debt = 0.0

    for sale in sales:
        total_debt += float(sale.debt_amount or 0)
        total_revenue += float(sale.total_amount or 0)
        if sale.status == "pagado":
            paid_count += 1
        for item in sale.items:
            product = item.product
            if not product:
                continue
            qty = item.quantity or 0
            total_cost += (product.cost_price or 0) * qty

    total_profit = total_revenue - total_cost
    profit_margin = (total_profit / total_revenue * 100) if total_revenue > 0 else 0
    sales_count = len(sales)

    return {
        "total_sales": sales_count,
        "paid_sales": paid_count,
        "pending_sales": sales_count - paid_count,
        "total_sales_amount": total_revenue,
        "total_cost_amount": total_cost,
        "total_profit": total_profit,
        "profit_margin": profit_margin,
        "total_debt": total_debt,
        "average_profit_per_sale": (total_profit / sales_count) if sales_count > 0 else 0
    }


def _top_products_for_period(sales: list, limit: int = 10) -> list:
    """Productos con mayor ganancia dentro de una lista de ventas ya filtrada."""
    products = {}
    for sale in sales:
        for item in sale.items:
            product = item.product
            if not product:
                continue
            qty = item.quantity or 0
            price = product.price or 0
            cost = product.cost_price or 0
            entry = products.setdefault(product.id, {
                "product_name": product.name,
                "total_quantity_sold": 0,
                "total_profit": 0.0,
                "sale_price": price
            })
            entry["total_quantity_sold"] += qty
            entry["total_profit"] += (price - cost) * qty

    return sorted(products.values(), key=lambda p: p["total_profit"], reverse=True)[:limit]


def generate_sales_excel(db: Session, company_id: int, start_date: datetime = None, end_date: datetime = None):
    """
    Genera un reporte en Excel con todas las ventas
    """
    if not PANDAS_AVAILABLE:
        raise Exception("pandas y openpyxl no están instalados")
    
    if not start_date:
        start_date = datetime.utcnow() - timedelta(days=30)
    if not end_date:
        end_date = datetime.utcnow()

    # Ventas de la compañía en el período. A diferencia del resto de la
    # tabla, acá SÍ se listan las ventas anuladas (para que el listado sirva
    # de auditoría), marcadas con estado "anulada"; el resumen de abajo no
    # las cuenta.
    sales = db.query(Sale).filter(
        Sale.company_id == company_id,
        Sale.created_at >= start_date,
        Sale.created_at <= end_date
    ).options(
        joinedload(Sale.items).joinedload(SaleItem.product),
        joinedload(Sale.customer)
    ).order_by(Sale.created_at.desc()).all()

    # Crear workbook con estilos
    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas"

    # Headers
    headers = ["ID Venta", "Cliente", "Fecha", "Total", "Pagado", "Deuda", "Estado", "Productos"]
    ws.append(headers)

    # Estilos
    header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Datos
    for sale in sales:
        estado = "anulada" if sale.cancelled_at else (sale.status or "pendiente")
        ws.append([
            sale.id,
            sale.customer.name if sale.customer else "Desconocido",
            sale.created_at.strftime("%Y-%m-%d %H:%M"),
            f"${sale.total_amount:.2f}",
            f"${sale.paid_amount:.2f}",
            f"${sale.debt_amount:.2f}",
            estado,
            len(sale.items)
        ])

    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12

    # Agregar resumen (solo ventas activas, no anuladas)
    active_sales = [s for s in sales if not s.cancelled_at]
    stats = _summarize_sales(active_sales)

    ws.append([])
    ws.append(["RESUMEN"])
    ws.append(["Total Ventas", stats["total_sales"]])
    ws.append(["Ventas Pagadas", stats["paid_sales"]])
    ws.append(["Ventas Pendientes", stats["pending_sales"]])
    ws.append(["Total Ingresos", f"${stats['total_sales_amount']:.2f}"])
    ws.append(["Costo Total", f"${stats['total_cost_amount']:.2f}"])
    ws.append(["Ganancia Total", f"${stats['total_profit']:.2f}"])
    ws.append(["Margen de Ganancia", f"{stats['profit_margin']:.1f}%"])
    ws.append(["Deuda Total", f"${stats['total_debt']:.2f}"])
    
    # Guardar a bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def generate_daily_report_pdf(db: Session, company_id: int, box_date: datetime = None):
    """
    Genera un reporte PDF del día (caja diaria)
    """
    if not REPORTLAB_AVAILABLE:
        raise Exception("reportlab no está instalado")

    if not box_date:
        box_date = datetime.utcnow().date()

    # Obtener caja del día
    daily_box = db.query(DailyBox).filter(
        DailyBox.date == box_date.isoformat(),
        DailyBox.company_id == company_id
    ).first()

    # Obtener ventas del día
    start_of_day = datetime.combine(box_date, datetime.min.time())
    end_of_day = datetime.combine(box_date, datetime.max.time())

    sales = _get_period_sales(db, company_id, start_of_day, end_of_day)
    stats = _summarize_sales(sales)
    
    # Crear PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1976D2'),
        spaceAfter=30,
        alignment=1  # Center
    )
    
    # Título
    elements.append(Paragraph("REPORTE DIARIO - CAJA", title_style))
    elements.append(Paragraph(f"Fecha: {box_date.strftime('%d/%m/%Y')}", styles['Normal']))
    elements.append(Spacer(1, 0.5*inch))
    
    # Tabla de resumen
    summary_data = [
        ["CONCEPTO", "VALOR"],
        ["Total Ventas", str(stats["total_sales"])],
        ["Ventas Pagadas", str(stats["paid_sales"])],
        ["Ventas Pendientes", str(stats["pending_sales"])],
        ["Total Ingresos", f"${stats['total_sales_amount']:.2f}"],
        ["Costo Total", f"${stats['total_cost_amount']:.2f}"],
        ["Ganancia Total", f"${stats['total_profit']:.2f}"],
        ["Margen de Ganancia", f"{stats['profit_margin']:.1f}%"],
        ["Deuda Generada", f"${stats['total_debt']:.2f}"],
    ]
    
    if daily_box:
        summary_data.append(["Apertura de Caja", f"${(daily_box.opening_balance or 0):.2f}"])
        if daily_box.closing_balance is not None:
            summary_data.append(["Cierre de Caja", f"${daily_box.closing_balance:.2f}"])

    table = Table(summary_data, colWidths=[3*inch, 2*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1976D2')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))

    elements.append(table)
    elements.append(Spacer(1, 0.5*inch))

    # Productos más vendidos
    elements.append(Paragraph("Productos Más Vendidos", styles['Heading2']))

    top_products = _top_products_for_period(sales, limit=5)
    
    if top_products:
        products_data = [["PRODUCTO", "CANTIDAD", "GANANCIA"]]
        for product in top_products:
            products_data.append([
                product["product_name"],
                str(product["total_quantity_sold"]),
                f"${product['total_profit']:.2f}"
            ])
        
        products_table = Table(products_data, colWidths=[3*inch, 1.5*inch, 1.5*inch])
        products_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1976D2')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(products_table)
    
    doc.build(elements)
    buffer.seek(0)
    
    return buffer


def generate_weekly_report_pdf(db: Session, company_id: int):
    """Reporte semanal"""
    end_date = datetime.utcnow()
    start_date = end_date - timedelta(days=7)

    return _generate_period_report_pdf(db, company_id, start_date, end_date, "SEMANAL")


def generate_monthly_report_pdf(db: Session, company_id: int):
    """Reporte mensual"""
    end_date = datetime.utcnow()
    start_date = end_date - timedelta(days=30)

    return _generate_period_report_pdf(db, company_id, start_date, end_date, "MENSUAL")


def _generate_period_report_pdf(db: Session, company_id: int, start_date: datetime, end_date: datetime, period_type: str):
    """Genera PDF para un período"""
    if not REPORTLAB_AVAILABLE:
        raise Exception("reportlab no está instalado")

    sales = _get_period_sales(db, company_id, start_date, end_date)
    stats = _summarize_sales(sales)
    top_products = _top_products_for_period(sales, limit=10)
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1976D2'),
        spaceAfter=30,
        alignment=1
    )
    
    # Título
    elements.append(Paragraph(f"REPORTE {period_type}", title_style))
    elements.append(Paragraph(
        f"Del {start_date.strftime('%d/%m/%Y')} al {end_date.strftime('%d/%m/%Y')}",
        styles['Normal']
    ))
    elements.append(Spacer(1, 0.5*inch))
    
    # Tabla principal
    summary_data = [
        ["CONCEPTO", "VALOR"],
        ["Total de Ventas", str(stats["total_sales"])],
        ["Ventas Pagadas", str(stats["paid_sales"])],
        ["Ventas Pendientes", str(stats["pending_sales"])],
        ["Total Ingresos", f"${stats['total_sales_amount']:.2f}"],
        ["Costo Total", f"${stats['total_cost_amount']:.2f}"],
        ["Ganancia REAL", f"${stats['total_profit']:.2f}"],
        ["Margen de Ganancia", f"{stats['profit_margin']:.1f}%"],
        ["Deuda Total", f"${stats['total_debt']:.2f}"],
        ["Ganancia Promedio/Venta", f"${stats['average_profit_per_sale']:.2f}"],
    ]
    
    table = Table(summary_data, colWidths=[3*inch, 2*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1976D2')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    elements.append(PageBreak())
    
    # Productos con mayor ganancia
    elements.append(Paragraph("Top 10 Productos por Ganancia", styles['Heading2']))
    elements.append(Spacer(1, 0.3*inch))
    
    if top_products:
        products_data = [["PRODUCTO", "VENDIDAS", "GANANCIA", "MARGEN"]]
        for product in top_products:
            margin = (product["total_profit"] / (product["sale_price"] * product["total_quantity_sold"]) * 100) if product["total_quantity_sold"] > 0 else 0
            products_data.append([
                product["product_name"],
                str(product["total_quantity_sold"]),
                f"${product['total_profit']:.2f}",
                f"{margin:.1f}%"
            ])
        
        products_table = Table(products_data, colWidths=[2.5*inch, 1*inch, 1.5*inch, 1*inch])
        products_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4CAF50')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
        ]))
        
        elements.append(products_table)
    
    doc.build(elements)
    buffer.seek(0)
    
    return buffer


# ─────────────────────────────────────────────────────────────────────────────
# EXPORTAR CAJA DIARIA — EXCEL
# ─────────────────────────────────────────────────────────────────────────────

def generate_daily_box_excel(db: Session, box_date_str: str, company_id: int):
    """
    Genera un Excel detallado de una caja diaria específica.
    Incluye: resumen de la caja, listado de ventas, desglose de productos.
    """
    if not PANDAS_AVAILABLE:
        raise Exception("openpyxl no está instalado")

    from app.models.daily_box import DailyBox
    from app.models.sale import Sale
    from app.models.sale_item import SaleItem
    from app.models.product import Product
    from sqlalchemy import and_
    from sqlalchemy.orm import joinedload

    box = db.query(DailyBox).filter(
        and_(DailyBox.date == box_date_str, DailyBox.company_id == company_id)
    ).first()

    date_start = datetime.strptime(box_date_str, "%Y-%m-%d")
    date_end = date_start.replace(hour=23, minute=59, second=59)

    if box:
        sales = db.query(Sale).filter(
            and_(Sale.daily_box_id == box.id, Sale.company_id == company_id, Sale.cancelled_at.is_(None))
        ).options(joinedload(Sale.items).joinedload(SaleItem.product), joinedload(Sale.customer)).all()
    else:
        sales = db.query(Sale).filter(
            and_(Sale.created_at >= date_start, Sale.created_at <= date_end, Sale.company_id == company_id, Sale.cancelled_at.is_(None))
        ).options(joinedload(Sale.items).joinedload(SaleItem.product), joinedload(Sale.customer)).all()

    wb = Workbook()

    # ── Hoja 1: Resumen ──────────────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Resumen Caja"

    DARK_BLUE  = "0F172A"
    MID_BLUE   = "1E3A5F"
    ACCENT     = "3B82F6"
    SUCCESS    = "22C55E"
    DANGER     = "EF4444"
    LIGHT_ROW  = "F0F4FF"
    WHITE      = "FFFFFF"
    GRAY       = "94A3B8"

    def h_font(bold=True, color=WHITE, size=11):
        return Font(bold=bold, color=color, size=size, name="Calibri")

    def fill(hex_color):
        return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

    def border():
        thin = Side(style="thin", color="CBD5E1")
        return Border(left=thin, right=thin, top=thin, bottom=thin)

    ws_summary.column_dimensions["A"].width = 30
    ws_summary.column_dimensions["B"].width = 22

    # Título
    ws_summary.merge_cells("A1:B1")
    ws_summary["A1"] = f"REPORTE DE CAJA DIARIA — {date_start.strftime('%d/%m/%Y')}"
    ws_summary["A1"].font = h_font(size=14)
    ws_summary["A1"].fill = fill(DARK_BLUE)
    ws_summary["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_summary.row_dimensions[1].height = 30

    if box:
        status_text = "ABIERTA" if box.status == "open" else "CERRADA"
        box_rows = [
            ("Estado de caja", status_text),
            ("Apertura", box.opened_at.strftime("%H:%M") if box.opened_at else "—"),
            ("Cierre", box.closed_at.strftime("%H:%M") if box.closed_at else "—"),
            ("Monto inicial", f"${box.opening_balance or 0:.2f}"),
            ("Monto final", f"${box.closing_balance or 0:.2f}" if box.closing_balance else "—"),
        ]
        for i, (label, value) in enumerate(box_rows, start=2):
            ws_summary[f"A{i}"] = label
            ws_summary[f"B{i}"] = value
            ws_summary[f"A{i}"].font = h_font(color="1E293B")
            ws_summary[f"B{i}"].font = h_font(bold=False, color="1E293B")
            bg = LIGHT_ROW if i % 2 == 0 else WHITE
            ws_summary[f"A{i}"].fill = fill(bg)
            ws_summary[f"B{i}"].fill = fill(bg)
            for col in ["A", "B"]:
                ws_summary[f"{col}{i}"].border = border()
        row_offset = 2 + len(box_rows) + 1
    else:
        row_offset = 3

    # Métricas de ventas
    total_sales_amt = sum(s.total_amount or 0 for s in sales)
    total_paid = sum(s.paid_amount or 0 for s in sales)
    total_debt = sum(s.debt_amount or 0 for s in sales)
    total_cost_of_sales = 0.0
    for sale in sales:
        for item in sale.items:
            p = item.product
            if p:
                total_cost_of_sales += (p.cost_price or 0) * (item.quantity or 0)
    total_profit = total_sales_amt - total_cost_of_sales

    ws_summary[f"A{row_offset}"] = "MÉTRICAS DEL DÍA"
    ws_summary.merge_cells(f"A{row_offset}:B{row_offset}")
    ws_summary[f"A{row_offset}"].font = h_font(size=12)
    ws_summary[f"A{row_offset}"].fill = fill(MID_BLUE)
    ws_summary[f"A{row_offset}"].alignment = Alignment(horizontal="center")
    row_offset += 1

    metrics = [
        ("Cantidad de ventas", str(len(sales))),
        ("Total facturado", f"${total_sales_amt:.2f}"),
        ("Total cobrado", f"${total_paid:.2f}"),
        ("Deuda generada", f"${total_debt:.2f}"),
        ("Ganancia del día", f"${total_profit:.2f}"),
        ("Margen", f"{(total_profit/total_sales_amt*100):.1f}%" if total_sales_amt > 0 else "—"),
    ]
    for i, (label, value) in enumerate(metrics):
        r = row_offset + i
        ws_summary[f"A{r}"] = label
        ws_summary[f"B{r}"] = value
        ws_summary[f"A{r}"].font = h_font(color="1E293B")
        ws_summary[f"B{r}"].font = h_font(bold=True, color="1E293B")
        bg = LIGHT_ROW if i % 2 == 0 else WHITE
        for col in ["A", "B"]:
            ws_summary[f"{col}{r}"].fill = fill(bg)
            ws_summary[f"{col}{r}"].border = border()

    # ── Hoja 2: Ventas detalladas ────────────────────────────────────────────
    ws_sales = wb.create_sheet("Ventas del Día")
    ws_sales.column_dimensions["A"].width = 10
    ws_sales.column_dimensions["B"].width = 22
    ws_sales.column_dimensions["C"].width = 14
    ws_sales.column_dimensions["D"].width = 14
    ws_sales.column_dimensions["E"].width = 14
    ws_sales.column_dimensions["F"].width = 14
    ws_sales.column_dimensions["G"].width = 12

    headers = ["#Factura", "Cliente", "Hora", "Total", "Pagado", "Deuda", "Productos"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws_sales.cell(row=1, column=col_idx, value=h)
        cell.font = h_font()
        cell.fill = fill(ACCENT)
        cell.alignment = Alignment(horizontal="center")
        cell.border = border()

    for row_idx, sale in enumerate(sales, start=2):
        vals = [
            f"#{sale.id:06d}",
            sale.customer.name if sale.customer else "Sin cliente",
            sale.created_at.strftime("%H:%M") if sale.created_at else "—",
            f"${sale.total_amount:.2f}",
            f"${sale.paid_amount:.2f}",
            f"${sale.debt_amount:.2f}",
            len(sale.items)
        ]
        bg = LIGHT_ROW if row_idx % 2 == 0 else WHITE
        for col_idx, val in enumerate(vals, start=1):
            cell = ws_sales.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill(bg)
            cell.alignment = Alignment(horizontal="center")
            cell.border = border()
            cell.font = Font(name="Calibri", color="1E293B")

    # ── Hoja 3: Productos vendidos ───────────────────────────────────────────
    ws_prod = wb.create_sheet("Productos Vendidos")
    ws_prod.column_dimensions["A"].width = 28
    ws_prod.column_dimensions["B"].width = 12
    ws_prod.column_dimensions["C"].width = 14
    ws_prod.column_dimensions["D"].width = 14
    ws_prod.column_dimensions["E"].width = 14

    prod_headers = ["Producto", "Unidades", "Ingreso", "Costo", "Ganancia"]
    for col_idx, h in enumerate(prod_headers, start=1):
        cell = ws_prod.cell(row=1, column=col_idx, value=h)
        cell.font = h_font()
        cell.fill = fill(SUCCESS)
        cell.alignment = Alignment(horizontal="center")
        cell.border = border()

    prod_totals = {}
    for sale in sales:
        for item in sale.items:
            p = item.product
            if not p:
                continue
            qty = item.quantity or 0
            price = p.price or 0
            cost = p.cost_price or 0
            if p.name not in prod_totals:
                prod_totals[p.name] = {"qty": 0, "revenue": 0, "cost": 0, "profit": 0}
            prod_totals[p.name]["qty"] += qty
            prod_totals[p.name]["revenue"] += price * qty
            prod_totals[p.name]["cost"] += cost * qty
            prod_totals[p.name]["profit"] += (price - cost) * qty

    sorted_prods = sorted(prod_totals.items(), key=lambda x: x[1]["profit"], reverse=True)
    for row_idx, (name, t) in enumerate(sorted_prods, start=2):
        vals = [name, t["qty"], f"${t['revenue']:.2f}", f"${t['cost']:.2f}", f"${t['profit']:.2f}"]
        bg = LIGHT_ROW if row_idx % 2 == 0 else WHITE
        for col_idx, val in enumerate(vals, start=1):
            cell = ws_prod.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill(bg)
            cell.alignment = Alignment(horizontal="center")
            cell.border = border()
            cell.font = Font(name="Calibri", color="1E293B")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ─────────────────────────────────────────────────────────────────────────────
# EXPORTAR CAJA DIARIA — PDF
# ─────────────────────────────────────────────────────────────────────────────

def generate_daily_box_pdf(db: Session, box_date_str: str, company_id: int):
    """
    Genera un PDF de la caja diaria con resumen, ventas y productos.
    """
    if not REPORTLAB_AVAILABLE:
        raise Exception("reportlab no está instalado")

    from app.models.daily_box import DailyBox
    from app.models.sale import Sale
    from app.models.sale_item import SaleItem
    from sqlalchemy import and_
    from sqlalchemy.orm import joinedload

    box = db.query(DailyBox).filter(
        and_(DailyBox.date == box_date_str, DailyBox.company_id == company_id)
    ).first()

    date_start = datetime.strptime(box_date_str, "%Y-%m-%d")
    date_end = date_start.replace(hour=23, minute=59, second=59)

    if box:
        sales = db.query(Sale).filter(
            and_(Sale.daily_box_id == box.id, Sale.company_id == company_id, Sale.cancelled_at.is_(None))
        ).options(joinedload(Sale.items).joinedload(SaleItem.product), joinedload(Sale.customer)).all()
    else:
        sales = db.query(Sale).filter(
            and_(Sale.created_at >= date_start, Sale.created_at <= date_end, Sale.company_id == company_id, Sale.cancelled_at.is_(None))
        ).options(joinedload(Sale.items).joinedload(SaleItem.product), joinedload(Sale.customer)).all()

    total_sales_amt = sum(s.total_amount or 0 for s in sales)
    total_paid = sum(s.paid_amount or 0 for s in sales)
    total_debt = sum(s.debt_amount or 0 for s in sales)
    total_cost_of_sales = 0.0
    prod_totals = {}
    for sale in sales:
        for item in sale.items:
            p = item.product
            if not p:
                continue
            qty = item.quantity or 0
            price = p.price or 0
            cost = p.cost_price or 0
            total_cost_of_sales += cost * qty
            if p.name not in prod_totals:
                prod_totals[p.name] = {"qty": 0, "revenue": 0, "profit": 0}
            prod_totals[p.name]["qty"] += qty
            prod_totals[p.name]["revenue"] += price * qty
            prod_totals[p.name]["profit"] += (price - cost) * qty
    total_profit = total_sales_amt - total_cost_of_sales

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            leftMargin=0.75*inch, rightMargin=0.75*inch,
                            topMargin=0.75*inch, bottomMargin=0.75*inch)
    elements = []
    styles = getSampleStyleSheet()

    BLUE = colors.HexColor("#1E3A5F")
    ACCENT = colors.HexColor("#3B82F6")
    LIGHT = colors.HexColor("#F0F4FF")
    GREEN = colors.HexColor("#22C55E")
    RED = colors.HexColor("#EF4444")

    title_style = ParagraphStyle("title", fontSize=18, textColor=BLUE,
                                 fontName="Helvetica-Bold", spaceAfter=4, alignment=1)
    sub_style = ParagraphStyle("sub", fontSize=10, textColor=colors.HexColor("#64748B"),
                               alignment=1, spaceAfter=16)
    section_style = ParagraphStyle("section", fontSize=12, textColor=BLUE,
                                   fontName="Helvetica-Bold", spaceBefore=14, spaceAfter=6)

    elements.append(Paragraph("REPORTE DE CAJA DIARIA", title_style))
    elements.append(Paragraph(date_start.strftime("%A, %d de %B de %Y").capitalize(), sub_style))

    # Resumen
    elements.append(Paragraph("Resumen del Día", section_style))
    summary_rows = [
        ["Concepto", "Valor"],
        ["Cantidad de ventas", str(len(sales))],
        ["Total facturado", f"${total_sales_amt:.2f}"],
        ["Total cobrado", f"${total_paid:.2f}"],
        ["Deuda generada", f"${total_debt:.2f}"],
        ["Ganancia neta", f"${total_profit:.2f}"],
        ["Margen de ganancia", f"{(total_profit/total_sales_amt*100):.1f}%" if total_sales_amt > 0 else "—"],
    ]
    if box:
        summary_rows += [
            ["Monto apertura", f"${box.opening_balance or 0:.2f}"],
            ["Monto cierre", f"${box.closing_balance or 0:.2f}" if box.closing_balance else "—"],
        ]

    t_summary = Table(summary_rows, colWidths=[3.5*inch, 2.5*inch])
    t_summary.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), ACCENT),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [LIGHT, colors.white]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    elements.append(t_summary)

    # Ventas
    if sales:
        elements.append(Paragraph("Detalle de Ventas", section_style))
        sale_rows = [["#Factura", "Cliente", "Hora", "Total", "Pagado", "Deuda"]]
        for s in sales:
            sale_rows.append([
                f"#{s.id:06d}",
                (s.customer.name if s.customer else "Sin cliente")[:22],
                s.created_at.strftime("%H:%M") if s.created_at else "—",
                f"${s.total_amount:.2f}",
                f"${s.paid_amount:.2f}",
                f"${s.debt_amount:.2f}",
            ])
        t_sales = Table(sale_rows, colWidths=[1*inch, 2*inch, 0.8*inch, 1.1*inch, 1.1*inch, 1*inch])
        t_sales.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), GREEN),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [LIGHT, colors.white]),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        elements.append(t_sales)

    # Productos
    if prod_totals:
        elements.append(Paragraph("Productos Vendidos", section_style))
        sorted_prods = sorted(prod_totals.items(), key=lambda x: x[1]["profit"], reverse=True)
        prod_rows = [["Producto", "Unidades", "Ingresos", "Ganancia"]]
        for name, t in sorted_prods:
            prod_rows.append([name[:30], str(t["qty"]), f"${t['revenue']:.2f}", f"${t['profit']:.2f}"])
        t_prod = Table(prod_rows, colWidths=[2.5*inch, 1*inch, 1.5*inch, 1.5*inch])
        t_prod.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), BLUE),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [LIGHT, colors.white]),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
            ("ALIGN", (1, 0), (-1, -1), "CENTER"),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        elements.append(t_prod)

    doc.build(elements)
    buffer.seek(0)
    return buffer

